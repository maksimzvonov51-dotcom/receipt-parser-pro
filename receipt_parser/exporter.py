import csv
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def export_to_csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    """
    Export parsed receipt rows to CSV bytes.
    """
    output = io.StringIO()

    if not rows:
        return b""

    fieldnames = list(rows[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames)

    writer.writeheader()
    writer.writerows(rows)

    return output.getvalue().encode("utf-8-sig")


def auto_adjust_columns(worksheet) -> None:
    """
    Adjust Excel column width based on cell content.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        worksheet.column_dimensions[column_letter].width = max_length + 2


def style_header_row(worksheet) -> None:
    """
    Apply simple business-style formatting to the first row.
    """
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def export_to_excel_bytes(
    rows: list[dict[str, Any]],
    summary: dict[str, Any] | None = None,
) -> bytes:
    """
    Export parsed receipt rows to an Excel file.

    The workbook contains:
    - Receipts: detailed receipt/invoice data
    - Summary: business summary dashboard data
    """
    workbook = Workbook()

    receipts_sheet = workbook.active
    receipts_sheet.title = "Receipts"

    if rows:
        headers = list(rows[0].keys())
        receipts_sheet.append(headers)

        for row in rows:
            receipts_sheet.append([row.get(header) for header in headers])

        style_header_row(receipts_sheet)
        auto_adjust_columns(receipts_sheet)

    else:
        receipts_sheet.append(["No data"])

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])

    if summary:
        summary_rows = [
            ["Total files", summary.get("total_files", 0)],
            ["Successfully parsed", summary.get("successfully_parsed", 0)],
            ["Failed", summary.get("failed", 0)],
            ["Total amount", summary.get("total_amount", 0)],
            ["Total VAT", summary.get("total_vat", 0)],
        ]

        for row in summary_rows:
            summary_sheet.append(row)
    else:
        summary_sheet.append(["No summary data", ""])

    style_header_row(summary_sheet)
    auto_adjust_columns(summary_sheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()